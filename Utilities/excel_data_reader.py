import openpyxl
import pandas as pd
from configuration.config_reader import get_config


def data_reader(sheet_name, column_name, excel_file_name):
    data = []
    file_path = f'C:/Users/manju/Downloads/{excel_file_name}.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    header_row = 1
    total_columns = sheet.max_column

    for col in range(1, total_columns + 1):
        if sheet.cell(header_row, col).value == column_name:
            column_number = col
            break
    else:
        raise ValueError(f"Column '{column_name}' not found in the sheet")

    total_rows = sheet.max_row
    for row in range(2, total_rows + 1):  # Assuming data starts from row 2
        data.append(sheet.cell(row, column_number).value)

    return data


def excel_to_dict(sheet_name, file_name):
    home_dir_file_path = get_config('path', 'table1')
    excel_file = home_dir_file_path + f"\\{file_name}.xlsx"
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    data_dict = df.to_dict(orient='list')

    return data_dict




