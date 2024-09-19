import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import os
import datetime
from configuration.config_reader import get_config


# Create the timestamped directory once and reuse it
def create_timestamped_directory():
    base_path = get_config("path", "home_dir")
    # Generate a timestamped directory name
    timestamp = datetime.datetime.now().strftime('data_validation_result_%Y-%m-%d_%H-%M-%S')
    timestamped_dir = os.path.join(base_path, timestamp)

    # Create the directory if it does not exist
    if not os.path.exists(timestamped_dir):
        os.makedirs(timestamped_dir)
        print(f"Created timestamped directory: {timestamped_dir}")

    return timestamped_dir


def apply_formatting(sheet, columns):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:  # Avoid applying borders to empty cells
                cell.border = thin_border
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width


# This function now takes 'timestamped_dir' as a parameter
def create_excel_sheet(columns, data, sheet_name, timestamped_dir):
    df = pd.DataFrame(data, columns=columns)
    excel_file_path = os.path.join(timestamped_dir, 'data_validation_results.xlsx')

    if os.path.exists(excel_file_path):
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
        except (KeyError, IOError) as e:
            print(f"Error loading workbook: {e}")
            workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.Workbook()

    # Create or update the sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # replace_or_append_table(sheet, data, columns)
    else:
        new_sheet = workbook.create_sheet(sheet_name)
        for col_idx, col in enumerate(columns):
            cell = new_sheet.cell(row=1, column=col_idx + 1)
            cell.value = col
            cell.font = Font(bold=True)  # Set header font to bold
        for row_idx, row in enumerate(data, start=2):
            for col_idx, value in enumerate(row):
                new_sheet.cell(row=row_idx, column=col_idx + 1).value = value
        print(f"New sheet '{sheet_name}' created in '{excel_file_path}'.")

    # Apply formatting to the sheet
    sheet = workbook[sheet_name]
    apply_formatting(sheet, columns)

    # Remove 'Sheet1' if it exists
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
        print(f"Sheet 'Sheet' has been deleted from '{excel_file_path}'.")

    workbook.save(excel_file_path)


# This function now takes 'timestamped_dir' as a parameter
def update_status_sheet(sheet_name, columns, timestamped_dir):
    excel_file_path = os.path.join(timestamped_dir, 'data_validation_results.xlsx')

    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    # Create or update the status sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Clear the sheet and add new headers
    sheet.delete_rows(1, sheet.max_row)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)

    # Gather the status data for each non-empty sheet
    sheet_names = [name for name in workbook.sheetnames if name != sheet_name]
    table_data = []
    for name in sheet_names:
        status, message = check_column_for_failed_values(name, 'Status', timestamped_dir)
        table_data.append([name, status, message])

    # Add the table data to the status sheet
    for row_idx, row_data in enumerate(table_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 2:
                if value.lower() == 'pass':
                    cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                elif value.lower() == 'fail':
                    cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    apply_formatting(sheet, columns)
    workbook.save(excel_file_path)
    print(f"Status summary updated in sheet '{sheet_name}' of '{excel_file_path}'.")


# This function now takes 'timestamped_dir' as a parameter
def check_column_for_failed_values(sheet_name, status_column, timestamped_dir):
    excel_file_path = os.path.join(timestamped_dir, 'data_validation_results.xlsx')

    workbook = openpyxl.load_workbook(excel_file_path)

    if sheet_name not in workbook.sheetnames:
        return "N/A", "Sheet not found"

    sheet = workbook[sheet_name]
    status_col_idx = None

    # Find the column with the specified status_column header
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
        if col[0].value == status_column:
            status_col_idx = col_idx + 1  # openpyxl is 1-indexed

    if status_col_idx is None:
        return "N/A", "Status column not found"

    # Check for any 'failed' or 'passed' values in the column
    status_list = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=status_col_idx, max_col=status_col_idx):
        cell_value = row[0].value
        if cell_value and cell_value.lower() == 'failed':
            status_list.append('failed')
        elif cell_value and cell_value.lower() == 'passed':
            status_list.append('passed')

    # Determine overall status
    if 'failed' in status_list:
        return 'fail', 'One or more rows has failed in this particular sheet'
    elif 'passed' in status_list:
        return 'pass', 'All rows have passed in this particular sheet'
    else:
        return 'N/A', 'No valid status values found'

